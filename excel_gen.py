"""
Generador de Excel con formato contable
"""
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generar_excel(movimientos, ruta_salida, empresa='', banco='', archivo=''):
    wb = Workbook()
    thin   = Side(border_style='thin',   color='BFBFBF')
    medium = Side(border_style='medium', color='7F7F7F')
    brd    = Border(left=thin, right=thin, top=thin, bottom=thin)
    brd_h  = Border(left=thin, right=thin, top=thin, bottom=medium)
    pf = '#,##0.00'
    C  = Alignment(horizontal='center', vertical='center')
    L  = Alignment(horizontal='left',   vertical='center', indent=1)
    R  = Alignment(horizontal='right',  vertical='center')
    AZH = PatternFill('solid', fgColor='1F3864')
    GP  = PatternFill('solid', fgColor='F5F5F5')
    BL  = PatternFill('solid', fgColor='FFFFFF')
    DF  = PatternFill('solid', fgColor='EBF3FB')
    HF  = PatternFill('solid', fgColor='F0F7EE')

    def H(ws, r, c, v):
        x = ws.cell(row=r, column=c, value=v)
        x.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        x.fill = AZH; x.alignment = C; x.border = brd; return x

    def D(ws, r, c, v, fill, fmt=None, aln=None, bold=False, col='000000'):
        x = ws.cell(row=r, column=c, value=v)
        x.font = Font(name='Arial', size=9, bold=bold, color=col)
        x.fill = fill; x.border = brd
        if fmt: x.number_format = fmt
        if aln: x.alignment = aln
        return x

    titulo = f"{empresa or 'Empresa'} — {banco.title() or 'Banco'} | {archivo or ''}"
    subtit = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    # Hoja 1: Movimientos
    ws1 = wb.active; ws1.title = 'Movimientos'
    ws1.merge_cells('A1:J1'); ws1['A1'] = titulo
    ws1['A1'].font = Font(name='Arial', bold=True, size=13, color='1F3864'); ws1['A1'].alignment = C
    ws1.row_dimensions[1].height = 22
    ws1.merge_cells('A2:J2'); ws1['A2'] = subtit
    ws1['A2'].font = Font(name='Arial', size=10, color='595959'); ws1['A2'].alignment = C
    ws1.row_dimensions[3].height = 4
    for c, h in enumerate(['Fecha','Comprobante','Descripción','Tipo','Débito ($)','Crédito ($)','Saldo ($)','Categoría','Cta. Debe','Cta. Haber'], 1):
        H(ws1, 4, c, h)
    ws1.row_dimensions[4].height = 17
    for i, m in enumerate(movimientos):
        r = i + 5; fill = BL if i%2==0 else GP; ec = m.get('tipo') == 'Crédito'
        D(ws1,r,1,m.get('fecha',''),fill,aln=C)
        D(ws1,r,2,m.get('comprobante',''),fill,aln=C)
        D(ws1,r,3,m.get('descripcion',''),fill,aln=L)
        D(ws1,r,4,m.get('tipo',''),fill,aln=C,bold=True,col='375623' if ec else '843C0C')
        D(ws1,r,5,m.get('debito'),fill,fmt=pf,aln=R)
        D(ws1,r,6,m.get('credito'),fill,fmt=pf,aln=R)
        D(ws1,r,7,m.get('saldo'),fill,fmt=pf,aln=R)
        D(ws1,r,8,m.get('categoria',''),fill,aln=L)
        D(ws1,r,9,m.get('cuenta_debe',''),fill,aln=L)
        D(ws1,r,10,m.get('cuenta_haber',''),fill,aln=L)
        ws1.row_dimensions[r].height = 13
    for w, col in zip([11,13,58,10,14,14,14,26,24,24], 'ABCDEFGHIJ'):
        ws1.column_dimensions[col].width = w
    ws1.freeze_panes = 'A5'

    # Hoja 2: Asientos
    ws2 = wb.create_sheet('Asientos contables')
    ws2.merge_cells('A1:F1'); ws2['A1'] = f'Libro Diario — {titulo}'
    ws2['A1'].font = Font(name='Arial', bold=True, size=13, color='1F3864'); ws2['A1'].alignment = C
    ws2.row_dimensions[1].height = 22; ws2.row_dimensions[2].height = 4
    for c, h in enumerate(['N° Asiento','Fecha','Cuenta','Glosa','Debe ($)','Haber ($)'], 1): H(ws2, 3, c, h)
    ws2.row_dimensions[3].height = 17
    cur = 4; num = 1
    for m in movimientos:
        imp = m.get('credito') if m.get('tipo') == 'Crédito' else m.get('debito')
        if not imp: continue
        ws2.cell(row=cur,column=1,value=num).font=Font(name='Arial',size=9,bold=True,color='1F3864')
        ws2.cell(row=cur,column=1).fill=DF;ws2.cell(row=cur,column=1).alignment=C;ws2.cell(row=cur,column=1).border=brd
        ws2.cell(row=cur,column=2,value=m.get('fecha','')).font=Font(name='Arial',size=9)
        ws2.cell(row=cur,column=2).fill=DF;ws2.cell(row=cur,column=2).alignment=C;ws2.cell(row=cur,column=2).border=brd
        c3=ws2.cell(row=cur,column=3,value=m.get('cuenta_debe',''));c3.font=Font(name='Arial',size=9,bold=True,color='1F3864')
        c3.fill=DF;c3.alignment=L;c3.border=brd
        c4=ws2.cell(row=cur,column=4,value=m.get('descripcion',''));c4.font=Font(name='Arial',size=9)
        c4.fill=DF;c4.alignment=L;c4.border=brd
        c5=ws2.cell(row=cur,column=5,value=imp);c5.font=Font(name='Arial',size=9,bold=True)
        c5.fill=DF;c5.number_format=pf;c5.alignment=R;c5.border=brd
        ws2.cell(row=cur,column=6).fill=DF;ws2.cell(row=cur,column=6).border=brd
        ws2.row_dimensions[cur].height=13;cur+=1
        ws2.cell(row=cur,column=1).fill=HF;ws2.cell(row=cur,column=1).border=brd
        ws2.cell(row=cur,column=2).fill=HF;ws2.cell(row=cur,column=2).border=brd
        c3h=ws2.cell(row=cur,column=3,value=f'    {m.get("cuenta_haber","")}');c3h.font=Font(name='Arial',size=9,bold=True,color='2E7D32')
        c3h.fill=HF;c3h.alignment=L;c3h.border=brd_h
        c4h=ws2.cell(row=cur,column=4,value=f'a / {m.get("descripcion","")}');c4h.font=Font(name='Arial',size=9,color='404040',italic=True)
        c4h.fill=HF;c4h.alignment=L;c4h.border=brd_h
        ws2.cell(row=cur,column=5).fill=HF;ws2.cell(row=cur,column=5).border=brd_h
        c6h=ws2.cell(row=cur,column=6,value=imp);c6h.font=Font(name='Arial',size=9,bold=True,color='2E7D32')
        c6h.fill=HF;c6h.number_format=pf;c6h.alignment=R;c6h.border=brd_h
        ws2.row_dimensions[cur].height=13;cur+=1;num+=1
    tot = sum((m.get('debito') or 0)+(m.get('credito') or 0) for m in movimientos)
    ws2.row_dimensions[cur].height=17;ws2.merge_cells(f'A{cur}:D{cur}')
    ct=ws2.cell(row=cur,column=1,value='TOTALES — Debe = Haber')
    ct.font=Font(name='Arial',bold=True,color='FFFFFF',size=10);ct.fill=AZH;ct.alignment=R;ct.border=brd
    for col,val in [(5,tot),(6,tot)]:
        c=ws2.cell(row=cur,column=col,value=val)
        c.font=Font(name='Arial',bold=True,color='FFFFFF',size=10)
        c.fill=AZH;c.number_format=pf;c.alignment=R;c.border=brd
    for w,col in zip([12,11,28,56,18,18],'ABCDEF'): ws2.column_dimensions[col].width=w
    ws2.freeze_panes='A4'

    # Hoja 3: Resumen
    ws3 = wb.create_sheet('Resumen por categoría')
    for c,h in enumerate(['Categoría','Movimientos','Débitos ($)','Créditos ($)','Neto ($)'],1): H(ws3,1,c,h)
    cats = {}
    for m in movimientos:
        cat = m.get('categoria','Sin clasificar')
        if cat not in cats: cats[cat]={'n':0,'deb':0.,'cred':0.}
        cats[cat]['n']+=1; cats[cat]['deb']+=m.get('debito') or 0; cats[cat]['cred']+=m.get('credito') or 0
    for i,(cat,d) in enumerate(sorted(cats.items()),2):
        fill=BL if i%2==0 else GP; neto=d['cred']-d['deb']
        D(ws3,i,1,cat,fill,aln=L);D(ws3,i,2,d['n'],fill,aln=C)
        D(ws3,i,3,d['deb'] or None,fill,fmt=pf,aln=R);D(ws3,i,4,d['cred'] or None,fill,fmt=pf,aln=R)
        D(ws3,i,5,neto,fill,fmt=pf,aln=R,bold=True,col='375623' if neto>=0 else '843C0C')
        ws3.row_dimensions[i].height=13
    for w,col in zip([32,12,18,18,18],'ABCDE'): ws3.column_dimensions[col].width=w

    wb.save(ruta_salida)
